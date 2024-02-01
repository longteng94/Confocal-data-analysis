# -*- coding: utf-8 -*-
"""
Created on Wed Apr 19 13:48:54 2023

@author: iSheratonX
"""

import numpy as np
from numpy import ndarray
import pandas as pd
import openpyxl

# Read the maxtri of txt
file_extension = np.loadtxt(r'F:\!Pro slide electrification affect the charged stuff adsorption\DATA\QD\QD_negtive drop\QD-COOH 10 µL drop on PFOTS.lif - 83 nM - C=1.txt')

matrix = file_extension

# create empty lists to store the x, y, and z values
x_values = []
y_values = []
z_values = []

# loop through each row and column of the matrix and append the x, y, and z values to the corresponding lists
for x in range(len(matrix)):
    for y in range(len(matrix[x])):
        x_values.append(x)
        y_values.append(y)
        z_values.append(matrix[x][y])

# combine the x, y, and z lists into a single list of tuples
data = list(zip(x_values, y_values, z_values))

# print the data as a table with headers
print("{:<10} {:<10} {:<10}".format('x', 'y', 'z'))
for x, y, z in data:
    print("{:<10} {:<10} {:<10}".format(x, y, z))
    
# create a new Excel workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# add headers to the first row of the worksheet
ws['A1'] = 'x'
ws['B1'] = 'y'
ws['C1'] = 'z'

# loop through each row of data and write the values to the worksheet
for row_num, (x, y, z) in enumerate(data, start=2):
    ws.cell(row=row_num, column=1, value=x)
    ws.cell(row=row_num, column=2, value=y)
    ws.cell(row=row_num, column=3, value=z)

# save the workbook to a file
wb.save('QD-COOH 10 µL drop on PFOTS.lif - 83 nM - C=1.xlsx')


